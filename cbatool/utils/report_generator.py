"""
Standardized Report Generator for CBAtool v2.0

Provides a flexible reporting mechanism for cable burial analysis.
"""

import os
import logging
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Union, Tuple
from datetime import datetime

# Configure logging
logger = logging.getLogger(__name__)

# Add openpyxl imports for formatting
try:
	import openpyxl
	from openpyxl.styles import PatternFill, Font, Alignment
	OPENPYXL_AVAILABLE = True
except ImportError:
	OPENPYXL_AVAILABLE = False
	logger.warning("openpyxl styles not available - Excel formatting will be limited")

try:
	from reportlab.lib.pagesizes import A4
	from reportlab.lib import colors
	from reportlab.lib.units import inch
	from reportlab.lib.units import cm
	from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
	from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
	REPORTLAB_AVAILABLE = True
except ImportError:
	REPORTLAB_AVAILABLE = False
	logger.warning("ReportLab is not installed. PDF generation will not be available.")

class ReportGenerator:
	"""
	Centralized report generation class for cable analysis.
	
	Works with any analyzer type (depth, position, combined) to produce
	consistent reports in multiple formats (Excel, PDF, HTML visualization).
	
	Attributes:
		output_directory (str): Directory where reports will be saved
		report_paths (Dict[str, str]): Dictionary of generated report paths
		report_config (Dict[str, Any]): Configuration for report generation
	"""
	
	def __init__(self, output_directory: str, report_config: Optional[Dict[str, Any]] = None):
		"""
		Initialize the ReportGenerator.
		
		Args:
			output_directory: Directory where reports will be saved
			report_config: Optional configuration dictionary for report generation
		"""
		self.output_directory = output_directory
		
		# Ensure output directory exists
		os.makedirs(output_directory, exist_ok=True)
		
		# Initialize report paths dictionary
		self.report_paths = {}
		
		# Set default report configuration if none provided
		self.report_config = report_config or {
			'include_anomalies': True,
			'include_problem_sections': True,
			'include_summary': True,
			'include_recommendations': True,
			'company_name': '',
			'project_name': '',
			'client_name': '',
			'regulatory_requirements': ''
		}
		
		# Logging setup
		logger.info(f"ReportGenerator initialized. Output directory: {output_directory}")
	
	def generate_reports(self, analysis_results: Dict[str, Any], 
						 visualization_path: Optional[str] = None,
						 analysis_type: str = 'combined') -> Dict[str, str]:
		"""
		Generate a complete set of reports from analysis results.
		
		This is the main entry point for report generation. It will produce:
		- Excel report with comprehensive data
		- PDF summary report
		- Links to visualization files
		
		Args:
			analysis_results: Dictionary containing analysis results from any analyzer
			visualization_path: Optional path to HTML visualization file
			analysis_type: Type of analysis ('depth', 'position', or 'combined')
			
		Returns:
			Dictionary of generated report paths
		"""
		# Validate input
		if not analysis_results:
			logger.warning("No analysis results provided for report generation")
			return {}
			
		# Normalize analysis results to standard format
		standardized_results = self._standardize_analysis_results(analysis_results, analysis_type)
		
		# Generate individual Excel reports
		excel_reports = self._generate_excel_reports(standardized_results, analysis_type)
		
		# Consolidate Excel reports if we have any
		consolidated_excel_path = ""
		if excel_reports:
			consolidated_excel_path = self.consolidate_excel_reports(
				excel_reports, 
				f"{analysis_type}_analysis_report.xlsx"
			)
		
		# Generate PDF summary
		pdf_path = self.generate_pdf_summary(
			standardized_results, 
			visualization_path,
			f"{analysis_type}_analysis_summary.pdf"
		)
		
		# Collect and return all report paths
		report_paths = {
			'excel': consolidated_excel_path,
			'pdf': pdf_path,
			'visualization': visualization_path,
			'individual_reports': excel_reports
		}
		
		# Store report paths for later reference
		self.report_paths = report_paths
		
		return report_paths
	
	def _standardize_analysis_results(self, analysis_results: Dict[str, Any], 
							 analysis_type: str) -> Dict[str, Any]:
		"""
		Standardize analysis results to a consistent format.
		
		This ensures that regardless of analyzer type, we have a consistent
		structure for report generation.
		
		Args:
			analysis_results: Raw analysis results from an analyzer
			analysis_type: Type of analysis ('depth', 'position', or 'combined')
			
		Returns:
			Standardized analysis results dictionary
		"""
		standardized = {
			'analysis_type': analysis_type,
			'timestamp': datetime.now(),
			'summary': {},
			'problem_sections': {},
			'anomalies': {},
			'recommendations': []
		}
		
		try:
			# Log what we received for debugging purposes
			logger.info(f"Standardizing results for analysis_type: {analysis_type}")
			if isinstance(analysis_results, dict):
				logger.info(f"Analysis results keys: {list(analysis_results.keys())}")
			else:
				logger.info(f"Analysis results is of type: {type(analysis_results)}")
			
			# Check if we're dealing with an analyzer object or a results dictionary
			if hasattr(analysis_results, 'get_analysis_summary') and callable(getattr(analysis_results, 'get_analysis_summary')):
				# It's an analyzer object
				logger.info("Processing analyzer object")
				summary = analysis_results.get_analysis_summary()
				if summary:
					standardized['summary'] = summary
					
				# Try to get problem sections and anomalies from analyzer
				if hasattr(analysis_results, 'analysis_results'):
					analyzer_results = analysis_results.analysis_results
					self._extract_from_analyzer_results(analyzer_results, standardized, analysis_type)
			else:
				# It's a dictionary
				logger.info("Processing dictionary results")
				# Extract common fields from analysis_results
				if isinstance(analysis_results, dict):
					if 'analysis_complete' in analysis_results:
						standardized['analysis_complete'] = analysis_results['analysis_complete']
					
					# Process depth analysis results
					if analysis_type in ['depth', 'combined']:
						self._extract_depth_results(analysis_results, standardized)
						
					# Process position analysis results
					if analysis_type in ['position', 'combined']:
						self._extract_position_results(analysis_results, standardized)
			
			# Generate recommendations based on analysis results
			standardized['recommendations'] = self._generate_recommendations(standardized)
			
			# Log what we've extracted to help with debugging
			logger.info(f"Standardized results summary keys: {list(standardized['summary'].keys())}")
			logger.info(f"Problem sections: {list(standardized['problem_sections'].keys())}")
			logger.info(f"Anomalies: {list(standardized['anomalies'].keys())}")
			
			return standardized
		except Exception as e:
			logger.error(f"Error in _standardize_analysis_results: {str(e)}", exc_info=True)
			return standardized
			
	def _extract_from_analyzer_results(self, analyzer_results, standardized, analysis_type):
		"""Extract data from analyzer results object."""
		try:
			# Extract problem sections
			if 'problem_sections' in analyzer_results:
				problems = analyzer_results['problem_sections']
				if isinstance(problems, pd.DataFrame) and not problems.empty:
					standardized['problem_sections'][analysis_type] = problems
					logger.info(f"Extracted problem sections for {analysis_type}")
			
			# Extract anomalies
			if 'anomalies' in analyzer_results:
				anomalies = analyzer_results['anomalies']
				if isinstance(anomalies, pd.DataFrame) and not anomalies.empty:
					standardized['anomalies'][analysis_type] = anomalies
					logger.info(f"Extracted anomalies for {analysis_type}")
			
			# Extract additional summary metrics
			for key in ['compliance_percentage', 'target_depth', 'section_count', 'total_problem_length']:
				if key in analyzer_results:
					standardized['summary'][f"{analysis_type}_{key}"] = analyzer_results[key]
		except Exception as e:
			logger.error(f"Error extracting from analyzer results: {str(e)}", exc_info=True)
	
	def _extract_depth_results(self, analysis_results: Dict[str, Any], 
							  standardized: Dict[str, Any]) -> None:
		"""
		Extract and standardize depth analysis results.
		
		Args:
			analysis_results: Raw analysis results
			standardized: Standardized results to update
		"""
		# Extract summary information
		if 'compliance_percentage' in analysis_results:
			standardized['summary']['depth_compliance_percentage'] = analysis_results['compliance_percentage']
			
		if 'target_depth' in analysis_results:
			standardized['summary']['target_depth'] = analysis_results['target_depth']
			
		# Extract problem sections
		if 'problem_sections' in analysis_results:
			problem_sections = analysis_results['problem_sections']
			if isinstance(problem_sections, pd.DataFrame) and not problem_sections.empty:
				standardized['problem_sections']['depth'] = problem_sections
				
				# Extract counts by severity if available
				if 'section_count' in analysis_results:
					standardized['summary']['depth_problem_section_count'] = analysis_results['section_count']
					
				if 'total_problem_length' in analysis_results:
					standardized['summary']['depth_total_problem_length'] = analysis_results['total_problem_length']
		
		# Extract anomalies
		if 'anomalies' in analysis_results:
			anomalies = analysis_results['anomalies']
			if isinstance(anomalies, pd.DataFrame) and not anomalies.empty:
				standardized['anomalies']['depth'] = anomalies
				
				# Add anomaly count to summary
				standardized['summary']['depth_anomaly_count'] = len(anomalies)
	
	def _extract_position_results(self, analysis_results: Dict[str, Any], 
								 standardized: Dict[str, Any]) -> None:
		"""
		Extract and standardize position analysis results.
		
		Args:
			analysis_results: Raw analysis results
			standardized: Standardized results to update
		"""
		try:
			# Check for position_analysis section in combined results
			position_results = analysis_results
			if 'position_analysis' in analysis_results:
				position_results = analysis_results['position_analysis']
			elif self.report_config.get('analysis_type') != 'position':
				logger.info("No position_analysis key found in results")
				return
				
			logger.info(f"Processing position results: {type(position_results)}")
			if isinstance(position_results, dict):
				logger.info(f"Position results keys: {list(position_results.keys())}")
	  
			# Try both problem_sections and problem_segments keys (for compatibility)
			problem_key = None
			for key in ['problem_sections', 'problem_segments']:
				if key in position_results:
					problem_key = key
					break
					
			if problem_key:
				problem_data = position_results.get(problem_key)
				# Handle problem sections data if it exists and is not empty
				if problem_data is not None and isinstance(problem_data, pd.DataFrame) and not problem_data.empty:
					standardized['problem_sections']['position'] = problem_data
					logger.info(f"Extracted position problem data from {problem_key}")
			
			# Extract summary information from position results or nested summary
			position_summary = position_results.get('summary', {})
			if position_summary:
				# Add relevant fields to standardized summary
				for key, value in position_summary.items():
					if not isinstance(value, dict):  # Skip nested dictionaries
						standardized['summary'][f'position_{key}'] = value
				
				# Handle nested anomalies dictionary
				if 'anomalies' in position_summary:
					standardized['summary']['position_anomalies'] = position_summary['anomalies']
			
			# Extract position anomalies from data
			position_data = position_results.get('position_analysis', None)
			if position_data is None and isinstance(position_results, pd.DataFrame):
				position_data = position_results
				
			if position_data is not None and isinstance(position_data, pd.DataFrame):
				# Filter anomalous points based on common position anomaly flags
				anomaly_flags = [
					'Is_KP_Jump', 'Is_KP_Reversal', 'Is_KP_Duplicate', 
					'Is_Significant_Deviation'
				]
				
				# Check which flags exist in the DataFrame
				existing_flags = [flag for flag in anomaly_flags if flag in position_data.columns]
				
				if existing_flags:
					# Create a combined filter for any anomaly
					anomaly_filter = False
					for flag in existing_flags:
						anomaly_filter |= position_data[flag]
					
					# Extract anomalous points
					position_anomalies = position_data[anomaly_filter].copy()
					
					if not position_anomalies.empty:
						standardized['anomalies']['position'] = position_anomalies
						standardized['summary']['position_anomaly_count'] = len(position_anomalies)
						logger.info(f"Extracted {len(position_anomalies)} position anomalies")
		except Exception as e:
			logger.error(f"Error extracting position results: {str(e)}", exc_info=True)
	
	def _generate_recommendations(self, standardized_results: Dict[str, Any]) -> List[Dict[str, Any]]:
		"""
		Generate recommendations based on standardized analysis results.
		
		Args:
			standardized_results: Standardized analysis results
			
		Returns:
			List of recommendation dictionaries
		"""
		recommendations = []
		
		# Helper function to add a recommendation
		def add_recommendation(category, severity, description, action):
			recommendations.append({
				'category': category,
				'severity': severity,
				'description': description,
				'action': action
			})
		
		# Depth analysis recommendations
		if 'depth' in standardized_results.get('problem_sections', {}):
			depth_sections = standardized_results['problem_sections']['depth']
			
			if isinstance(depth_sections, pd.DataFrame) and not depth_sections.empty and 'Severity' in depth_sections.columns:
				# Check for high severity sections
				high_severity = depth_sections[depth_sections['Severity'] == 'High']
				if not high_severity.empty:
					add_recommendation(
						'Depth', 'High',
						f"Found {len(high_severity)} high severity burial depth issues",
						"Remedial burial required for these sections"
					)
				
				# Check for medium severity sections
				medium_severity = depth_sections[depth_sections['Severity'] == 'Medium']
				if not medium_severity.empty:
					add_recommendation(
						'Depth', 'Medium',
						f"Found {len(medium_severity)} medium severity burial depth issues",
						"Consider additional protection for these sections"
					)
				
				# Check for low severity sections
				low_severity = depth_sections[depth_sections['Severity'] == 'Low']
				if not low_severity.empty:
					add_recommendation(
						'Depth', 'Low',
						f"Found {len(low_severity)} low severity burial depth issues",
						"Monitor these sections during maintenance"
					)
		
		# Position analysis recommendations
		if 'position' in standardized_results.get('problem_sections', {}):
			position_sections = standardized_results['problem_sections']['position']
			
			if isinstance(position_sections, pd.DataFrame) and not position_sections.empty:
				# Check if severity is available
				if 'Severity' in position_sections.columns:
					# Generate recommendations by severity
					for severity in ['High', 'Medium', 'Low']:
						severity_sections = position_sections[position_sections['Severity'] == severity]
						if not severity_sections.empty:
							add_recommendation(
								'Position', severity,
								f"Found {len(severity_sections)} {severity.lower()} severity position issues",
								"Review position data quality in these sections"
							)
				else:
					# Generic recommendation if severity not available
					add_recommendation(
						'Position', 'Medium',
						f"Found {len(position_sections)} position quality issues",
						"Review position data quality in problem sections"
					)
		
		# Check for specific position anomaly types
		position_anomalies = standardized_results.get('summary', {}).get('position_anomalies', {})
		if position_anomalies:
			if position_anomalies.get('kp_jumps', 0) > 0:
				add_recommendation(
					'Position', 'Medium',
					f"Found {position_anomalies['kp_jumps']} KP jumps",
					"Investigate KP continuity issues"
				)
			
			if position_anomalies.get('kp_reversals', 0) > 0:
				add_recommendation(
					'Position', 'High',
					f"Found {position_anomalies['kp_reversals']} KP reversals",
					"Review position data sequence and direction"
				)
		
		return recommendations
		
	def _generate_excel_reports(self, standardized_results: Dict[str, Any], 
							   analysis_type: str) -> Dict[str, str]:
		"""
		Generate individual Excel reports from standardized analysis results.
		
		Args:
			standardized_results: Standardized analysis results
			analysis_type: Type of analysis ('depth', 'position', or 'combined')
			
		Returns:
			Dictionary mapping report names to file paths
		"""
		excel_reports = {}
		
		# Generate problem sections reports
		for section_type, sections_df in standardized_results.get('problem_sections', {}).items():
			if isinstance(sections_df, pd.DataFrame) and not sections_df.empty:
				report_name = f"{section_type}_problem_sections"
				file_path = os.path.join(self.output_directory, f"{report_name}.xlsx")
				
				# Save to Excel
				sections_df.to_excel(file_path, index=False)
				
				# Add to report dictionary
				excel_reports[f"{section_type.capitalize()} Problem Sections"] = file_path
		
		# Generate anomalies reports
		for anomaly_type, anomalies_df in standardized_results.get('anomalies', {}).items():
			if isinstance(anomalies_df, pd.DataFrame) and not anomalies_df.empty:
				report_name = f"{anomaly_type}_anomalies"
				file_path = os.path.join(self.output_directory, f"{report_name}.xlsx")
				
				# Save to Excel
				anomalies_df.to_excel(file_path, index=False)
				
				# Add to report dictionary
				excel_reports[f"{anomaly_type.capitalize()} Anomalies"] = file_path
		
		# Generate recommendations report
		recommendations = standardized_results.get('recommendations', [])
		if recommendations:
			report_name = "recommendations"
			file_path = os.path.join(self.output_directory, f"{report_name}.xlsx")
			
			# Convert to DataFrame
			recommendations_df = pd.DataFrame(recommendations)
			
			# Save to Excel
			recommendations_df.to_excel(file_path, index=False)
			
			# Add to report dictionary
			excel_reports["Recommendations"] = file_path
		
		# Generate summary report
		summary = standardized_results.get('summary', {})
		if summary:
			report_name = "analysis_summary"
			file_path = os.path.join(self.output_directory, f"{report_name}.xlsx")
			
			# Convert to DataFrame (two columns: Metric and Value)
			summary_rows = []
			for key, value in summary.items():
				if not isinstance(value, dict):  # Skip nested dictionaries
					summary_rows.append({
						'Metric': key.replace('_', ' ').title(),
						'Value': value
					})
			
			summary_df = pd.DataFrame(summary_rows)
			
			# Save to Excel
			summary_df.to_excel(file_path, index=False)
			
			# Add to report dictionary
			excel_reports["Analysis Summary"] = file_path
		
		return excel_reports
	
	def consolidate_excel_reports(self, 
								 reports: Dict[str, str], 
								 output_filename: str = 'comprehensive_analysis_report.xlsx') -> str:
		"""
		Consolidate multiple Excel reports into a single workbook.
		
		Args:
			reports: Dictionary mapping report names to file paths
			output_filename: Name of the output consolidated report
		
		Returns:
			Path to the consolidated Excel report
		"""
		# Validate input
		if not reports:
			logger.warning("No reports provided for consolidation")
			return ""
		
		# Full output path
		output_path = os.path.join(self.output_directory, output_filename)
		
		try:
			# Create Excel writer
			with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
				# First add a summary sheet
				summary_data = self._create_summary_sheet(reports)
				summary_data.to_excel(writer, sheet_name='Summary', index=False)
				
				# Add recommendations sheet if it exists
				if "Recommendations" in reports:
					try:
						recommendations_df = pd.read_excel(reports["Recommendations"])
						recommendations_df.to_excel(writer, sheet_name='Recommendations', index=False)
						
						# Apply formatting to recommendations
						self._apply_excel_formatting(writer, 'Recommendations', recommendations_df, 
												   format_type='recommendations')
					except Exception as e:
						logger.error(f"Error adding recommendations: {e}")
				
				# Add each report
				for report_name, report_path in reports.items():
					# Skip recommendations as we've already added it
					if report_name == "Recommendations":
						continue
						
					# Validate report exists
					if not os.path.exists(report_path):
						logger.warning(f"Report not found: {report_path}")
						continue
					
					# Read Excel file
					try:
						xls = pd.ExcelFile(report_path)
						
						# Add each sheet to the consolidated workbook
						for sheet_name in xls.sheet_names:
							# Create a better sheet name
							friendly_sheet_name = f"{report_name}"
							if len(friendly_sheet_name) > 31:  # Excel sheet name length limit
								friendly_sheet_name = friendly_sheet_name[:31]
							
							# Read and write sheet
							df = pd.read_excel(report_path, sheet_name=sheet_name)
							df.to_excel(writer, sheet_name=friendly_sheet_name, index=False)
							
							# Apply formatting if possible
							self._apply_excel_formatting(writer, friendly_sheet_name, df)
					except Exception as e:
						logger.error(f"Error processing report {report_name}: {e}")
				
				# Add project information sheet
				self._add_project_info_sheet(writer)
			
			logger.info(f"Consolidated report created: {output_path}")
			self.report_paths['excel'] = output_path
			return output_path
		
		except Exception as e:
			logger.error(f"Error consolidating reports: {e}")
			return ""

	def generate_pdf_summary(self, standardized_results: Dict[str, Any], 
                        visualization_path: Optional[str] = None,
                        output_filename: str = 'analysis_summary.pdf') -> str:
		"""
		Generate a PDF summary report from analysis results.
		
		Args:
			standardized_results: Standardized analysis results
			visualization_path: Optional path to visualization file
			output_filename: Name of the output PDF file
			
		Returns:
			Path to the generated PDF file or empty string if generation failed
		"""
		# Import utility function
		from ..utils.report_utils import extract_severity_distribution
		
		# Skip PDF generation if ReportLab is not available
		if not REPORTLAB_AVAILABLE:
			logger.warning("ReportLab not available - skipping PDF generation")
			return ""
		
		output_path = os.path.join(self.output_directory, output_filename)
		logger.info(f"Generating PDF summary report at: {output_path}")
		
		try:
			# Validate visualization path if provided
			if visualization_path and not os.path.exists(visualization_path):
				logger.warning(f"Visualization file not found: {visualization_path}")
				visualization_path = None
			
			# Create PDF document
			doc = SimpleDocTemplate(
				output_path,
				pagesize=A4,
				rightMargin=72,
				leftMargin=72,
				topMargin=72,
				bottomMargin=72
			)
			
			# Initialize story (content elements)
			story = []
			
			# Get styles
			styles = getSampleStyleSheet()
			title_style = styles['Title']
			heading_style = styles['Heading1']
			subheading_style = styles['Heading2']
			normal_style = styles['Normal']
			
			# Add custom style for section headers
			section_style = ParagraphStyle(
				'SectionStyle',
				parent=styles['Heading2'],
				backColor=colors.lightgrey,
				borderPadding=5,
			)
			
			# Add custom style for recommendations
			rec_style = ParagraphStyle(
				'RecommendationStyle',
				parent=styles['Normal'],
				fontSize=10,
				leftIndent=20,
				firstLineIndent=0,
				spaceAfter=6
			)
			
			# Add title
			project_name = self.report_config.get('project_name', 'Cable Analysis Report')
			story.append(Paragraph(project_name, title_style))
			story.append(Spacer(1, 0.25 * inch))
			
			# Add report info
			now = datetime.now().strftime("%Y-%m-%d %H:%M")
			story.append(Paragraph(f"Report Generated: {now}", normal_style))
			
			if self.report_config.get('client_name'):
				story.append(Paragraph(f"Client: {self.report_config['client_name']}", normal_style))
				
			if self.report_config.get('company_name'):
				story.append(Paragraph(f"Company: {self.report_config['company_name']}", normal_style))
			
			story.append(Spacer(1, 0.25 * inch))
			
			# Add analysis type section
			analysis_type = standardized_results.get('analysis_type', 'combined').title()
			story.append(Paragraph(f"{analysis_type} Analysis Summary", heading_style))
			story.append(Spacer(1, 0.1 * inch))
			
			summary = standardized_results.get('summary', {})
			logger.info(f"PDF generation: Summary has {len(summary)} entries")
			
			# Create summary table data
			summary_data = [['Metric', 'Value']]
			
			for key, value in summary.items():
				if isinstance(value, dict):
					continue  # Skip nested dictionaries
				
				# Format the key for display
				display_key = key.replace('_', ' ').title()
				
				# Format value based on type
				if isinstance(value, float):
					display_value = f"{value:.2f}"
				elif isinstance(value, tuple) and len(value) == 2:
					display_value = f"{value[0]} - {value[1]}"
				else:
					display_value = str(value)
					
				summary_data.append([display_key, display_value])
			
			# Create table
			if len(summary_data) > 1:
				summary_table = Table(summary_data, colWidths=[doc.width * 0.6, doc.width * 0.4])
				summary_table.setStyle(TableStyle([
					('BACKGROUND', (0, 0), (1, 0), colors.lightgrey),
					('TEXTCOLOR', (0, 0), (1, 0), colors.black),
					('ALIGN', (0, 0), (1, 0), 'CENTER'),
					('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
					('BOTTOMPADDING', (0, 0), (1, 0), 12),
					('BACKGROUND', (0, 1), (-1, -1), colors.white),
					('GRID', (0, 0), (-1, -1), 1, colors.black),
					('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
				]))
				story.append(summary_table)
				story.append(Spacer(1, 0.25 * inch))
			else:
				story.append(Paragraph("No summary data available.", normal_style))
				story.append(Spacer(1, 0.25 * inch))
			
			# Add problem sections if available
			problem_sections = standardized_results.get('problem_sections', {})
			if problem_sections:
				logger.info(f"PDF generation: Problem sections types: {type(problem_sections)}")
				story.append(Paragraph("Problem Sections", heading_style))
				story.append(Spacer(1, 0.1 * inch))
				
				# Get total count of problem sections
				total_count = 0
				if isinstance(problem_sections, dict) and 'total_count' in problem_sections:
					total_count = problem_sections['total_count']
				elif isinstance(problem_sections, dict) and 'details' in problem_sections:
					total_count = len(problem_sections['details'])
				elif isinstance(problem_sections, dict):
					# Old format with multiple section types
					for section_type, sections in problem_sections.items():
						if isinstance(sections, pd.DataFrame):
							total_count += len(sections)
				
				story.append(Paragraph(f"Found {total_count} problem sections.", normal_style))
				
				# Get severity distribution
				severity_text = extract_severity_distribution(standardized_results)
				if severity_text:
					story.append(Paragraph(f"Severity distribution: {severity_text}", normal_style))
				
				story.append(Spacer(1, 0.15 * inch))
			
			# Add recommendations if available
			recommendations = standardized_results.get('recommendations', [])
			if recommendations:
				story.append(Paragraph("Recommendations", heading_style))
				story.append(Spacer(1, 0.1 * inch))
				
				for i, rec in enumerate(recommendations):
					category = rec.get('category', 'General')
					severity = rec.get('severity', 'Medium')
					description = rec.get('description', '')
					action = rec.get('action', '')
					action_items = rec.get('action_items', [])
					
					# Create colored bullet based on severity
					if severity.lower() == 'high':
						bullet_color = colors.red
					elif severity.lower() == 'medium':
						bullet_color = colors.orange
					else:
						bullet_color = colors.green
					
					bullet = f'<font color="{bullet_color}">\u2022</font>'
					
					# Add recommendation with colored bullet
					story.append(Paragraph(f"{bullet} <b>{category} ({severity}):</b> {description}", normal_style))
					
					# Add action or action items
					if action:
						story.append(Paragraph(f"   Action: {action}", normal_style))
					
					if action_items:
						for item in action_items:
							story.append(Paragraph(f"   • {item}", normal_style))
					
					story.append(Spacer(1, 0.1 * inch))
			
			# Add information about visualization
			if visualization_path and os.path.exists(visualization_path):
				story.append(Paragraph("Visualization", heading_style))
				story.append(Spacer(1, 0.1 * inch))
				story.append(Paragraph(
					f"An interactive visualization is available at: {os.path.basename(visualization_path)}",
					normal_style
				))
			
			# Add regulatory requirements if specified
			if self.report_config.get('regulatory_requirements'):
				story.append(Spacer(1, 0.25 * inch))
				story.append(Paragraph("Regulatory Requirements", heading_style))
				story.append(Spacer(1, 0.1 * inch))
				story.append(Paragraph(self.report_config['regulatory_requirements'], normal_style))
			
			# Build the PDF
			doc.build(story)
			logger.info(f"PDF summary report created: {output_path}")
			return output_path
			
		except Exception as e:
			logger.error(f"Error generating PDF summary: {str(e)}", exc_info=True)
			return ""
			
			# Add report information
			for report_name, report_path in reports.items():
				if os.path.exists(report_path):
					try:
						xls = pd.ExcelFile(report_path)
						total_rows = 0
						
						# Count total rows across all sheets
						for sheet_name in xls.sheet_names:
							df = pd.read_excel(report_path, sheet_name=sheet_name)
							total_rows += len(df)
						
						summary_rows.append({
							'Report Type': report_name,
							'File': os.path.basename(report_path),
							'Sheets': len(xls.sheet_names),
							'Total Rows': total_rows,
							'Date Generated': datetime.fromtimestamp(os.path.getmtime(report_path)).strftime('%Y-%m-%d %H:%M')
						})
					except Exception as e:
						logger.error(f"Error analyzing report {report_name}: {e}")
						summary_rows.append({
							'Report Type': report_name,
							'File': os.path.basename(report_path),
							'Sheets': 'Error',
							'Total Rows': 'Error',
							'Date Generated': datetime.fromtimestamp(os.path.getmtime(report_path)).strftime('%Y-%m-%d %H:%M')
						})
			
			# Add timestamp and metadata
			metadata = [{
				'Report Type': 'METADATA',
				'File': 'N/A',
				'Sheets': 'N/A',
				'Total Rows': 'N/A',
				'Date Generated': datetime.now().strftime('%Y-%m-%d %H:%M')
			}]
			
			return pd.DataFrame(metadata + summary_rows)
	
	def _add_project_info_sheet(self, writer):
		"""
		Add a project information sheet to the Excel report.
		
		Args:
			writer: Excel writer object
		"""
		# Create project info
		project_info = [
			{'Field': 'Project Name', 'Value': self.report_config.get('project_name', '')},
			{'Field': 'Client Name', 'Value': self.report_config.get('client_name', '')},
			{'Field': 'Company Name', 'Value': self.report_config.get('company_name', '')},
			{'Field': 'Report Generated', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M')},
			{'Field': 'Regulatory Requirements', 'Value': self.report_config.get('regulatory_requirements', '')}
		]
		
		# Convert to DataFrame
		df = pd.DataFrame(project_info)
		
		# Write to Excel
		df.to_excel(writer, sheet_name='Project Information', index=False)
		
		# Apply formatting
		if OPENPYXL_AVAILABLE:
			try:
				worksheet = writer.sheets['Project Information']
				
				# Format headers
				for cell in worksheet["A1:B1"]:
					for c in cell:
						c.font = Font(bold=True)
						c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
				
				# Auto-fit columns
				for idx, col in enumerate(['A', 'B']):
					max_length = 0
					for cell in worksheet[col]:
						try:
							if len(str(cell.value)) > max_length:
								max_length = len(str(cell.value))
						except:
							pass
					adjusted_width = (max_length + 2)
					worksheet.column_dimensions[col].width = adjusted_width
			except Exception as e:
				logger.warning(f"Error formatting project info sheet: {e}")
	
	def _apply_excel_formatting(self, writer, sheet_name, df, format_type='default'):
		"""
		Apply formatting to Excel worksheet.
		
		Args:
			writer: Excel writer object
			sheet_name: Name of the sheet to format
			df: DataFrame containing the data
			format_type: Type of formatting to apply ('default', 'recommendations', etc.)
		"""
		try:
			# Only proceed if openpyxl is available
			if not OPENPYXL_AVAILABLE:
				return
				
			# Get the workbook and worksheet
			workbook = writer.book
			worksheet = writer.sheets[sheet_name]
			
			# Auto-fit columns if possible
			for idx, col in enumerate(df.columns):
				# Calculate max width
				max_len = max(
					df[col].astype(str).map(len).max(),
					len(str(col))
				) + 2  # Add a little extra space
				
				# Set column width (using Excel column letters A, B, C...)
				col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx // 26) + chr(65 + idx % 26)
				if hasattr(worksheet, 'column_dimensions'):
					worksheet.column_dimensions[col_letter].width = min(max_len, 40)
			
			# Format header row
			for cell in worksheet["1:1"]:
				cell.font = Font(bold=True)
				cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
			
			# Apply specific formatting based on format_type
			if format_type == 'recommendations':
				self._format_recommendations(worksheet, df)
			elif 'Severity' in df.columns:
				self._format_severity_column(worksheet, df)
			
		except Exception as e:
			# Don't let formatting errors prevent report generation
			logger.warning(f"Error applying Excel formatting: {e}")
	
	def _format_severity_column(self, worksheet, df):
		"""
		Format the Severity column with color coding.
		
		Args:
			worksheet: Excel worksheet
			df: DataFrame containing the data
		"""
		# Find the severity column
		severity_col_idx = list(df.columns).index('Severity')
		severity_col_letter = chr(65 + severity_col_idx) if severity_col_idx < 26 else chr(64 + severity_col_idx // 26) + chr(65 + severity_col_idx % 26)
		
		# Apply formatting to each cell in the severity column
		for row_idx, severity in enumerate(df['Severity'], start=2):  # Start from row 2 (after header)
			cell = worksheet[f"{severity_col_letter}{row_idx}"]
			
			if severity == 'High':
				cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
			elif severity == 'Medium':
				cell.fill = PatternFill(start_color='FFEECC', end_color='FFEECC', fill_type='solid')
			elif severity == 'Low':
				cell.fill = PatternFill(start_color='EEFFCC', end_color='EEFFCC', fill_type='solid')
	
	def _format_recommendations(self, worksheet, df):
		"""
		Format the recommendations sheet with color coding and proper alignment.
		
		Args:
			worksheet: Excel worksheet
			df: DataFrame containing the data
		"""
		# Check if 'severity' column exists
		if 'severity' in df.columns:
			severity_col_idx = list(df.columns).index('severity')
			severity_col_letter = chr(65 + severity_col_idx) if severity_col_idx < 26 else chr(64 + severity_col_idx // 26) + chr(65 + severity_col_idx % 26)
			
			# Apply formatting to each cell in the severity column
			for row_idx, severity in enumerate(df['severity'], start=2):  # Start from row 2 (after header)
				cell = worksheet[f"{severity_col_letter}{row_idx}"]
				
				if severity == 'High':
					cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
				elif severity == 'Medium':
					cell.fill = PatternFill(start_color='FFEECC', end_color='FFEECC', fill_type='solid')
				elif severity == 'Low':
					cell.fill = PatternFill(start_color='EEFFCC', end_color='EEFFCC', fill_type='solid')
				
		# Check if 'category' column exists
		if 'category' in df.columns:
			# Format category column
			category_col_idx = list(df.columns).index('category')
			category_col_letter = chr(65 + category_col_idx) if category_col_idx < 26 else chr(64 + category_col_idx // 26) + chr(65 + category_col_idx % 26)
			
			# Make category bold
			for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
				cell = worksheet[f"{category_col_letter}{row_idx}"]
				cell.font = Font(bold=True)
		
		# Check if 'action' column exists
		if 'action' in df.columns:
			# Apply word wrap to action column
			action_col_idx = list(df.columns).index('action')
			action_col_letter = chr(65 + action_col_idx) if action_col_idx < 26 else chr(64 + action_col_idx // 26) + chr(65 + action_col_idx % 26)
			
			for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
				cell = worksheet[f"{action_col_letter}{row_idx}"]
				cell.alignment = Alignment(wrap_text=True)
	
	# Targeted fix for the create_comprehensive_report method
	def create_comprehensive_report(self, 
								analyzer_results: Dict[str, Any], 
								visualization_path: Optional[str] = None) -> Dict[str, str]:
		"""
		Create a comprehensive report from analyzer results.
		
		Args:
			analyzer_results: Dictionary containing analyzer results or an Analyzer instance
			visualization_path: Optional path to visualization file
		
		Returns:
			Dictionary of generated report paths
		"""
		# Import the report utility functions
		from ..utils.report_utils import (
			validate_standardized_results,
			extract_problem_sections,
			extract_anomalies,
			extract_recommendations,
			create_compliance_metrics_dataframe
		)
		
		# Step 1: Get standardized results
		standardized_data = None
		
		# Check if we're dealing with an analyzer object with the new API
		if hasattr(analyzer_results, 'get_standardized_results') and callable(getattr(analyzer_results, 'get_standardized_results')):
			logger.info("Using get_standardized_results method from analyzer")
			standardized_data = analyzer_results.get_standardized_results()
		else:
			# Determine if it's already standardized
			if isinstance(analyzer_results, dict) and 'analysis_type' in analyzer_results:
				logger.info("Using provided data as standardized results")
				standardized_data = analyzer_results
			else:
				# Use the old standardization method
				analysis_type = 'combined'
				if isinstance(analyzer_results, dict):
					if 'depth_analysis' in analyzer_results and 'position_analysis' not in analyzer_results:
						analysis_type = 'depth'
					elif 'position_analysis' in analyzer_results and 'depth_analysis' not in analyzer_results:
						analysis_type = 'position'
						
				logger.info(f"Standardizing analyzer results for {analysis_type} analysis")
				standardized_data = self._standardize_analysis_results(analyzer_results, analysis_type)
		
		# Step 2: Validate standardized results
		if not validate_standardized_results(standardized_data):
			logger.error("Invalid standardized data structure")
			return {}
		
		# Step 3: Check visualization path
		if visualization_path and not os.path.exists(visualization_path):
			logger.warning(f"Visualization file not found: {visualization_path}")
			visualization_path = None
		
		# Step 4: Get analysis type for consistent filenames
		analysis_type = standardized_data.get('analysis_type', 'combined')
		
		# Step 5: Create individual Excel reports based on extracted data
		report_files = {}
		
		# Process problem sections
		problem_sections_df = extract_problem_sections(standardized_data)
		if problem_sections_df is not None and not problem_sections_df.empty:
			problem_sections_path = os.path.join(self.output_directory, f"{analysis_type}_problem_sections.xlsx")
			problem_sections_df.to_excel(problem_sections_path, index=False)
			report_files[f'{analysis_type.capitalize()} Problem Sections'] = problem_sections_path
			logger.info(f"Created problem sections report: {problem_sections_path}")
		
		# Process anomalies
		anomalies_df = extract_anomalies(standardized_data)
		if anomalies_df is not None and not anomalies_df.empty:
			anomalies_path = os.path.join(self.output_directory, f"{analysis_type}_anomalies.xlsx")
			anomalies_df.to_excel(anomalies_path, index=False)
			report_files[f'{analysis_type.capitalize()} Anomalies'] = anomalies_path
			logger.info(f"Created anomalies report: {anomalies_path}")
		
		# Process recommendations
		recommendations_df = extract_recommendations(standardized_data)
		if recommendations_df is not None and not recommendations_df.empty:
			recommendations_path = os.path.join(self.output_directory, "recommendations.xlsx")
			recommendations_df.to_excel(recommendations_path, index=False)
			report_files['Recommendations'] = recommendations_path
			logger.info(f"Created recommendations report: {recommendations_path}")
		
		# Process compliance metrics
		compliance_df = create_compliance_metrics_dataframe(standardized_data)
		if compliance_df is not None and not compliance_df.empty:
			compliance_path = os.path.join(self.output_directory, "compliance_metrics.xlsx")
			compliance_df.to_excel(compliance_path, index=False)
			report_files['Compliance Metrics'] = compliance_path
			logger.info(f"Created compliance metrics report: {compliance_path}")
		
		# Step 6: Generate consolidated Excel report
		excel_report = self.consolidate_excel_reports(report_files, f"{analysis_type}_analysis_report.xlsx")
		
		# Step 7: Generate PDF summary
		pdf_report = self.generate_pdf_summary(
			standardized_data, 
			visualization_path,
			f"{analysis_type}_analysis_summary.pdf"
		)
		
		# Return all generated reports
		return {
			'excel_report': excel_report,
			'pdf_report': pdf_report,
			'visualization': visualization_path,
			'individual_reports': report_files
		}
	