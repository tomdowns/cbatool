
"""
Utility functions for report generation in CBAtool v2.0.

This module contains helper functions to support the ReportGenerator class.
"""

import pandas as pd
from typing import Dict, Any, List, Optional
from datetime import datetime
import logging

# Configure logging
logger = logging.getLogger(__name__)

def validate_standardized_results(data: Dict[str, Any]) -> bool:
    """
    Validate that the results follow the standardized structure.
    
    Args:
        data: The data to validate
        
    Returns:
        True if data follows the standard structure, False otherwise
    """
    # Basic structure check
    if not isinstance(data, dict):
        logger.warning("Data is not a dictionary")
        return False
    
    # Check for analysis_type
    if 'analysis_type' not in data:
        logger.warning("Missing analysis_type in standardized data")
        return False
        
    # Structure is valid enough to proceed
    return True

def extract_problem_sections(data: Dict[str, Any]) -> Optional[pd.DataFrame]:
    """
    Extract problem sections from standardized data structure.
    
    Works with both old format (DataFrame) and new format (details list).
    
    Args:
        data: Standardized data structure
        
    Returns:
        DataFrame containing problem sections or None if not available
    """
    if 'problem_sections' not in data:
        return None
    
    # Handle new standardized format
    if isinstance(data['problem_sections'], dict) and 'details' in data['problem_sections']:
        details = data['problem_sections']['details']
        if details:
            return pd.DataFrame(details)
        return None
    
    # Handle old format (direct DataFrame or dictionary of DataFrames)
    ps = data['problem_sections']
    
    # If it's already a DataFrame
    if isinstance(ps, pd.DataFrame):
        return ps if not ps.empty else None
        
    # If it's a dictionary of DataFrames
    if isinstance(ps, dict):
        for section_type, sections_df in ps.items():
            if isinstance(sections_df, pd.DataFrame) and not sections_df.empty:
                # Return the first non-empty DataFrame found
                # In the future, could combine them if needed
                return sections_df
    
    return None

def extract_anomalies(data: Dict[str, Any]) -> Optional[pd.DataFrame]:
    """
    Extract anomalies from standardized data structure.
    
    Works with both old format (DataFrame) and new format (details list).
    
    Args:
        data: Standardized data structure
        
    Returns:
        DataFrame containing anomalies or None if not available
    """
    if 'anomalies' not in data:
        return None
    
    # Handle new standardized format
    if isinstance(data['anomalies'], dict) and 'details' in data['anomalies']:
        details = data['anomalies']['details']
        if details:
            return pd.DataFrame(details)
        return None
    
    # Handle old format (direct DataFrame or dictionary of DataFrames)
    anom = data['anomalies']
    
    # If it's already a DataFrame
    if isinstance(anom, pd.DataFrame):
        return anom if not anom.empty else None
        
    # If it's a dictionary of DataFrames
    if isinstance(anom, dict):
        for anomaly_type, anomaly_df in anom.items():
            if isinstance(anomaly_df, pd.DataFrame) and not anomaly_df.empty:
                # Return the first non-empty DataFrame found
                return anomaly_df
    
    return None

def extract_recommendations(data: Dict[str, Any]) -> Optional[pd.DataFrame]:
    """
    Extract recommendations from standardized data structure.
    
    Args:
        data: Standardized data structure
        
    Returns:
        DataFrame containing recommendations or None if not available
    """
    if 'recommendations' not in data or not data['recommendations']:
        return None
    
    return pd.DataFrame(data['recommendations'])

def create_compliance_metrics_dataframe(data: Dict[str, Any]) -> Optional[pd.DataFrame]:
    """
    Create a DataFrame of compliance metrics from standardized data.
    
    Args:
        data: Standardized data structure
        
    Returns:
        DataFrame containing compliance metrics or None if not available
    """
    if 'compliance_metrics' not in data:
        return None
    
    metrics_rows = []
    
    # Process total compliance percentage
    total_compliance = data['compliance_metrics'].get('total_compliance_percentage')
    if total_compliance is not None:
        metrics_rows.append({
            'Metric': 'Total Compliance Percentage',
            'Value': f"{total_compliance:.2f}%"
        })
    
    # Process compliance by severity
    severity_metrics = data['compliance_metrics'].get('compliance_by_severity', {})
    for key, value in severity_metrics.items():
        if value is not None:
            metric_name = key.replace('_', ' ').title()
            metrics_rows.append({
                'Metric': metric_name,
                'Value': f"{value:.2f}%"
            })
    
    if metrics_rows:
        return pd.DataFrame(metrics_rows)
    return None

def extract_severity_distribution(data: Dict[str, Any], section_type: str = 'problem_sections') -> str:
		"""
		Extract formatted severity distribution from standardized data.
		
		Args:
			data: Standardized data structure
			section_type: Type of section ('problem_sections' or 'anomalies')
			
		Returns:
			Formatted string with severity distribution or empty string if not found
		"""
		# Handle new standardized format first
		if section_type in data and isinstance(data[section_type], dict) and 'severity_breakdown' in data[section_type]:
			breakdown = data[section_type]['severity_breakdown']
			severity_counts = []
			
			for severity, info in breakdown.items():
				count = info.get('count', 0)
				if count > 0:
					severity_counts.append(f"{count} {severity}")
			
			if severity_counts:
				return ", ".join(severity_counts)
		
		# Handle old format (direct DataFrames with 'Severity' column)
		if section_type in data:
			sections = data[section_type]
			
			# If it's a dictionary of DataFrames (old format)
			if isinstance(sections, dict):
				for _, df in sections.items():
					if isinstance(df, pd.DataFrame) and not df.empty and 'Severity' in df.columns:
						severity_counts = df['Severity'].value_counts().to_dict()
						formatted = [f"{count} {severity}" for severity, count in severity_counts.items()]
						if formatted:
							return ", ".join(formatted)
			
			# If it's a direct DataFrame
			elif isinstance(sections, pd.DataFrame) and not sections.empty and 'Severity' in sections.columns:
				severity_counts = sections['Severity'].value_counts().to_dict()
				formatted = [f"{count} {severity}" for severity, count in severity_counts.items()]
				if formatted:
					return ", ".join(formatted)
		
		return ""
def apply_excel_formatting(worksheet, df: pd.DataFrame, format_type: str = 'default'):
    """
    Apply standardized formatting to an Excel worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        df: DataFrame containing the data
        format_type: Type of formatting to apply ('default', 'recommendations', etc.)
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Auto-fit columns
        for idx, col in enumerate(df.columns):
            col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx // 26) + chr(65 + idx % 26)
            max_length = max([len(str(s)) for s in df[col].values] + [len(str(col))])
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[col_letter].width = min(adjusted_width, 40)
        
        # Format header row
        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply specific formatting based on format_type
        if format_type == 'recommendations':
            # Format severity column
            severity_col = None
            for col in df.columns:
                if col.lower() == 'severity':
                    severity_col = col
                    break
                    
            if severity_col:
                severity_col_idx = df.columns.get_loc(severity_col)
                severity_col_letter = chr(65 + severity_col_idx) if severity_col_idx < 26 else chr(64 + severity_col_idx // 26) + chr(65 + severity_col_idx % 26)
                
                # Apply formatting to each cell in the severity column
                for row_idx, severity in enumerate(df[severity_col], start=2):  # Start from row 2 (after header)
                    cell = worksheet[f"{severity_col_letter}{row_idx}"]
                    severity_value = str(severity).lower() if severity else ""
                    
                    if "high" in severity_value:
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    elif "medium" in severity_value:
                        cell.fill = PatternFill(start_color='FFEECC', end_color='FFEECC', fill_type='solid')
                    elif "low" in severity_value:
                        cell.fill = PatternFill(start_color='EEFFCC', end_color='EEFFCC', fill_type='solid')
            
            # Format action column with word wrap
            for col_name in ['action', 'action_items', 'description']:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name)
                    col_letter = chr(65 + col_idx) if col_idx < 26 else chr(64 + col_idx // 26) + chr(65 + col_idx % 26)
                    
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        elif format_type == 'problem_sections':
            # Format severity column
            severity_col = None
            for col in df.columns:
                if col.lower() == 'severity':
                    severity_col = col
                    break
                    
            if severity_col:
                severity_col_idx = df.columns.get_loc(severity_col)
                severity_col_letter = chr(65 + severity_col_idx) if severity_col_idx < 26 else chr(64 + severity_col_idx // 26) + chr(65 + severity_col_idx % 26)
                
                # Apply formatting to each cell in the severity column
                for row_idx, severity in enumerate(df[severity_col], start=2):  # Start from row 2 (after header)
                    cell = worksheet[f"{severity_col_letter}{row_idx}"]
                    severity_value = str(severity).lower() if severity else ""
                    
                    if "high" in severity_value:
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    elif "medium" in severity_value:
                        cell.fill = PatternFill(start_color='FFEECC', end_color='FFEECC', fill_type='solid')
                    elif "low" in severity_value:
                        cell.fill = PatternFill(start_color='EEFFCC', end_color='EEFFCC', fill_type='solid')
    
    except ImportError:
        # openpyxl not available
        pass
    except Exception as e:
        # Log but don't abort on formatting errors
        logging.warning(f"Excel formatting error: {str(e)}")